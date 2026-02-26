"""
Base layer-harvesting logic.

Handles JSON parsing, layer extraction, bbox detection, and Excel output.
No UI dependencies — import freely from any context.
"""

import sys
from pathlib import Path

import openpyxl  # noqa: F401  (imported for callers that do wb = openpyxl.Workbook())
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# When bundled with PyInstaller sys.frozen is set; use the directory containing
# the binary so envs/, input/, and output/ are always resolved relative to it,
# regardless of the working directory the binary is invoked from.
if getattr(sys, "frozen", False):
    _ROOT = Path(sys.executable).parent
else:
    _ROOT = Path(__file__).parent.parent

INPUT_DIR  = _ROOT / "input"
OUTPUT_DIR = _ROOT / "output"
ENVS_DIR   = _ROOT / "envs"

# ── Column definitions ────────────────────────────────────────────────────────

BASE_COLUMNS = [
    ("name",         "Layer Name"),
    ("title",        "Title"),
    ("abstract",     "Abstract"),
    ("queryable",    "Queryable"),
    ("crs",          "CRS"),
    ("west_bound",   "West Bound Lon"),
    ("east_bound",   "East Bound Lon"),
    ("north_bound",  "North Bound Lat"),
    ("south_bound",  "South Bound Lat"),
    ("is_global",    "Is Global"),
    ("style_names",  "Style Name(s)"),
    ("keyword_list", "Keywords"),
]

# PDF-related constants live here because extract_row() uses them
PDF_HAZARD_PREFIX = "pdf:hazardlookup:"
NO_PDF_LABEL      = "not attached to PDF V2"

# ── Bbox thresholds ───────────────────────────────────────────────────────────

GLOBAL_LON_THRESHOLD = 340
GLOBAL_LAT_THRESHOLD = 120

# ── Excel styles ──────────────────────────────────────────────────────────────

HEADER_FILL     = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT     = Font(color="FFFFFF", bold=True)
ROW_FILL_ODD    = PatternFill("solid", fgColor="D6E4F0")
ROW_FILL_EVEN   = PatternFill("solid", fgColor="FFFFFF")
ROW_FILL_GLOBAL = PatternFill("solid", fgColor="FFD700")

# ── Legend metadata ───────────────────────────────────────────────────────────

BASE_COLUMN_DESCRIPTIONS = [
    ("Layer Name",       "The WMS layer identifier (e.g. GRAPHRASTER:fires_final)."),
    ("Title",            "Human-readable display name of the layer."),
    ("Abstract",         "Brief description of the layer's content or purpose."),
    ("Queryable",        "1 = layer supports GetFeatureInfo requests; 0 = display-only."),
    ("CRS",              "Comma-separated list of supported coordinate reference systems."),
    ("West Bound Lon",   "Western edge of the bounding box in decimal degrees (−180 to 180)."),
    ("East Bound Lon",   "Eastern edge of the bounding box in decimal degrees (−180 to 180)."),
    ("North Bound Lat",  "Northern edge of the bounding box in decimal degrees (−90 to 90)."),
    ("South Bound Lat",  "Southern edge of the bounding box in decimal degrees (−90 to 90)."),
    ("Is Global",        f"'Yes' when the layer's bbox spans ≥ {GLOBAL_LON_THRESHOLD}° longitude "
                         f"AND ≥ {GLOBAL_LAT_THRESHOLD}° latitude, indicating worldwide coverage. "
                         f"'No' for regional or country-level layers."),
    ("Style Name(s)",    "Comma-separated WMS style names available for this layer."),
    ("Keywords",         "Full keyword_list from the capabilities document, comma-separated."),
]

ROW_HIGHLIGHTS = [
    (ROW_FILL_GLOBAL, "FFD700", "Global layer",  "Row highlighted in yellow — bbox qualifies as worldwide coverage (Is Global = Yes)."),
    (ROW_FILL_ODD,    "D6E4F0", "Odd data row",  "Light blue alternating row — regional or country-level layer (Is Global = No)."),
    (ROW_FILL_EVEN,   "FFFFFF", "Even data row", "White alternating row — regional or country-level layer (Is Global = No)."),
]


# ── Layer extraction ──────────────────────────────────────────────────────────

def find_hazard_layers(obj, results):
    """Walk the JSON tree iteratively and collect layers with 'hazardlookup' keyword."""
    stack = [obj]
    while stack:
        node = stack.pop()
        if isinstance(node, dict):
            kw = node.get("keyword_list") or []
            if isinstance(kw, list) and any("hazardlookup" in str(k).lower() for k in kw):
                results.append(node)
            else:
                stack.extend(node.values())
        elif isinstance(node, list):
            stack.extend(node)


def is_global_bbox(west, east, north, south) -> bool:
    """Return True when the bounding box covers nearly the entire world."""
    try:
        lon_span = float(east) - float(west)
        lat_span = float(north) - float(south)
        return lon_span >= GLOBAL_LON_THRESHOLD and lat_span >= GLOBAL_LAT_THRESHOLD
    except (TypeError, ValueError):
        return False


def extract_row(layer) -> dict:
    """Flatten a layer dict into all possible fields (usage-mode filtering happens in write_sheet)."""
    bbox = layer.get("ex_geographic_bounding_box") or {}
    if not isinstance(bbox, dict):
        bbox = {}

    west  = bbox.get("west_bound_longitude", "")
    east  = bbox.get("east_bound_longitude", "")
    north = bbox.get("north_bound_latitude", "")
    south = bbox.get("south_bound_latitude", "")

    styles = layer.get("style") or []
    style_names = ", ".join(
        s.get("name", "") for s in styles if isinstance(s, dict)
    ) if isinstance(styles, list) else ""

    crs = layer.get("CRS") or []
    crs_str = ", ".join(crs) if isinstance(crs, list) else str(crs)

    kw = layer.get("keyword_list") or []
    kw_str = ", ".join(str(k) for k in kw) if isinstance(kw, list) else str(kw)

    global_flag = is_global_bbox(west, east, north, south)

    # Always extracted; only written to the sheet when --usage pdf is active.
    # Normalise to lowercase so the extracted suffix is always lowercase regardless of source casing.
    pdf_v2 = next(
        (k.lower()[len(PDF_HAZARD_PREFIX):] for k in kw
         if isinstance(k, str) and k.lower().startswith(PDF_HAZARD_PREFIX)),
        NO_PDF_LABEL,
    )

    return {
        "name":         layer.get("name", ""),
        "title":        layer.get("title", ""),
        "abstract":     (layer.get("abstract") or "").strip(),
        "queryable":    layer.get("queryable", ""),
        "crs":          crs_str,
        "west_bound":   west,
        "east_bound":   east,
        "north_bound":  north,
        "south_bound":  south,
        "is_global":    "Yes" if global_flag else "No",
        "_global":      global_flag,
        "pdf_v2":       pdf_v2,
        "style_names":  style_names,
        "keyword_list": kw_str,
    }


# ── Excel sheet writing ───────────────────────────────────────────────────────

def write_sheet(ws, rows: list, columns: list):
    """Write header + data rows using only the given column list."""
    for col_idx, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row_data in enumerate(rows, start=2):
        if row_data.get("_global"):
            fill = ROW_FILL_GLOBAL
        elif row_idx % 2 == 0:
            fill = ROW_FILL_ODD
        else:
            fill = ROW_FILL_EVEN
        for col_idx, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="top")

    for col_idx, (key, label) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(label)
        for r in range(2, len(rows) + 2):
            val = ws.cell(row=r, column=col_idx).value or ""
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"


# ── Legend sheet ──────────────────────────────────────────────────────────────

def write_legend_sheet(wb, extra_col_desc=None):
    """
    Write the Legend sheet.

    extra_col_desc: optional (name, description) tuple inserted after the
                    'Is Global' entry (used by pdf to inject the PDF V2 row).
    """
    ws = wb.create_sheet(title="Legend")

    section_font = Font(bold=True, size=12)
    label_font   = Font(bold=True)
    wrap_align   = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Row Colour Key").font = section_font
    ws.cell(row=2, column=1, value="Colour").font    = label_font
    ws.cell(row=2, column=2, value="Label").font     = label_font
    ws.cell(row=2, column=3, value="Meaning").font   = label_font
    for r, (fill, hex_color, label, meaning) in enumerate(ROW_HIGHLIGHTS, start=3):
        swatch = ws.cell(row=r, column=1, value=f"  #{hex_color}  ")
        swatch.fill      = fill
        swatch.alignment = center_align
        ws.cell(row=r, column=2, value=label).font = Font(bold=False)
        ws.cell(row=r, column=3, value=meaning).alignment = wrap_align

    col_descs = list(BASE_COLUMN_DESCRIPTIONS)
    if extra_col_desc:
        is_global_idx = next((i for i, (n, _) in enumerate(col_descs) if n == "Is Global"), None)
        if is_global_idx is not None:
            col_descs.insert(is_global_idx + 1, extra_col_desc)
        else:
            col_descs.append(extra_col_desc)

    section_row = len(ROW_HIGHLIGHTS) + 5
    ws.cell(row=section_row, column=1, value="Column Descriptions").font = section_font
    header_row = section_row + 1
    ws.cell(row=header_row, column=1, value="Column").font      = label_font
    ws.cell(row=header_row, column=2, value="Description").font = label_font

    for i, (col_name, description) in enumerate(col_descs, start=header_row + 1):
        ws.cell(row=i, column=1, value=col_name).font = Font(bold=True)
        ws.cell(row=i, column=2, value=description).alignment = wrap_align

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 70


# ── Input scanning ────────────────────────────────────────────────────────────

def collect_groups(input_dir: Path) -> dict:
    """
    Scan input_dir and return an ordered dict:
      ""           → root-level *.json files  → output goes in OUTPUT_DIR
      "<subdir>"   → *.json inside that subdir → output goes in OUTPUT_DIR/<subdir>/
    Only direct subdirectories are considered (no deeper nesting).
    """
    if not input_dir.is_dir():
        return {}
    groups: dict[str, list[Path]] = {}
    root_files = sorted(input_dir.glob("*.json"))
    if root_files:
        groups[""] = root_files
    for subdir in sorted(d for d in input_dir.iterdir() if d.is_dir()):
        subdir_files = sorted(subdir.glob("*.json"))
        if subdir_files:
            groups[subdir.name] = subdir_files
    return groups
